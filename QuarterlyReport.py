#Isaiah Wallace 200557041
#A program which parses through two different spreadsheets for the needed info
#to make a performace report based on the info, then makes a diagnostic report
#to assure that all has been executed properly
import openpyxl, re, datetime, time
#Project and employee list paths
projectPath = "Info_For_Report_Cayman_Construction.xlsx"
employeePath = "Info_For_Report_Cayman_Employee_List_&_Hours.xlsx"
#patterns to look for in the spreadsheets
idPattern = r"\d\d\d\d\d\d"
employeeIdPattern = r"^\d{5}$"
clientRatingPattern = r"\d[.]\d" 
hoursPattern = r"^\d{3}$"
namePattern = r"^[a-zA-Z]+ [a-zA-Z]+$"
#arrays the scanners will put the relevant info into
idLog = []
clientRatingLog = []
employeeId = []
hoursLog = []
nameLog = []
employeeRating = []
performanceStatus = []
#columns names to use for the format of the generated speardsheet
reportColumnNames = ["Employee ID:","Name:","Project ID:","Hours:","Performance Status:"]
#searches throught the entire project sheet and collects info using regex patterns, storing
#info into arrays to build performance report
def projectScanner():
    #marks start time of program
    global startTime
    startTime = time.time()
    #counts the number of elements that aren't used by scanners
    global unusedElements
    unusedElements = 0
    print("Scanning Project Spreadsheet...")
    #loads spreadsheet and finds the max columns and rows
    project = openpyxl.load_workbook(projectPath)
    projectSheet = project.active
    projectColumns = projectSheet.max_column
    projectRows = projectSheet.max_row
    #searches through colums, starting at 1, as there isn't a column 0
    for i in range(1, projectColumns):
        projectCell = projectSheet.cell(row = 1, column = i)
        projectCellValue = projectCell.value
        #searches through rows, switching each column loop
        for j in range(1, projectRows):
            projectCellR = projectSheet.cell(row = j, column = i)
            #makes sure that if a None value is found, it is moved passed and doesn't
            #cause an exception
            try:
                #seacrhes through using pattern to find needed info
                if(re.search(idPattern, projectCellR.value)):
                    idLog.append(projectCellR.value)
                elif(re.search(clientRatingPattern, projectCellR.value)):
                     clientRatingLog.append(projectCellR.value) 
                else:
                    unusedElements+=1
                    
            #catches exception when a None value is found
            except TypeError:
                unusedElements+=1
    #returns completed when function is finished
    return "Completed"
#searches through employee spreadsheet, looking for needed info
def employeeListScanner():
    #counts the unused elements in the employee list
    global unusedElementsTwo
    unusedElementsTwo = 0
    print("Scanning Employee List Spreadsheet...")
    #looking through spreadsheet with path, making max colums and rows
    employee = openpyxl.load_workbook(employeePath)
    employeeSheet = employee.active
    employeeColumns = employeeSheet.max_column
    employeeRows = employeeSheet.max_row
    #looks through each column of employee spreadsheet
    for i in range(1, employeeColumns):
        employeeCell = employeeSheet.cell(row = 1, column = i)
        employeeCellValue = employeeCell.value
        #looks through the rows for all of the needed info
        for j in range(1, employeeRows):
            employeeCellR = employeeSheet.cell(row = j, column = i)
            try:
                #searches using name, hours and employee id pattern to find those values
                if(re.search(employeeIdPattern, employeeCellR.value)):
                    employeeId.append(employeeCellR.value)
                elif(re.search(hoursPattern, employeeCellR.value)):
                     hoursLog.append(employeeCellR.value)
                elif(re.search(namePattern, employeeCellR.value)):
                    nameLog.append(employeeCellR.value)
                else:
                    #adds to unused elements 
                    unusedElementsTwo+=1
            except TypeError:
                #adds to unused elements
                unusedElementsTwo+=1
    #shows that function has been completed
    return "Completed"
#generates spreadsheet report with info from the two searched through spreadsheets
def performanceReport():
    print("Generating Performance Report...")
    #creates a new spreadsheet
    workReport = openpyxl.Workbook()
    reportSheet = workReport.active
    #runs through the length of the found info, using a multipler picked based of of
    #the client rating then using it to find the employees overall score
    for i in range(len(hoursLog)):
        if(float(clientRatingLog[i]) >= 6.6 and float(clientRatingLog[i]) <= 10):
            employeeRatingMulti = 0.05
        elif(float(clientRatingLog[i]) >= 4 and float(clientRatingLog[i]) <= 6.5):
            employeeRatingMulti = 0.04
        elif(float(clientRatingLog[i]) >= 0 and float(clientRatingLog[i]) <= 3.9):
            employeeRatingMulti = 0.03
        rating = int(hoursLog[i])*employeeRatingMulti
        employeeRating.append(rating)
        #after finding the rating, using it to pick one of three statuses
        if(employeeRating[i] >= 6.6 and employeeRating[i] <= 10):
            performanceStatus.append("HIGH")
        elif(employeeRating[i] >= 4 and employeeRating[i] <= 6.5):
            performanceStatus.append("AVERAGE")
        elif(employeeRating[i] >= 0 and employeeRating[i] <= 3.9):
            performanceStatus.append("LOW")
    #for loops and arrays are zero indexed, columns and rows aren't so
    #another value is need to increment so all runs normally
    zeroIndexFix = 0
    #column number for each array of info
    columnNo = 1
    #runs through each column, each with it's own array to print to the columns
    for i in range(1,len(reportColumnNames)+1):
        #nested fix for zero indexed problem
        zeroIndexFixNested = 0
        reportSheetCell = reportSheet.cell(row = 1, column = i)
        reportSheetCell.value = reportColumnNames[zeroIndexFix]
        zeroIndexFix +=1
        #prints each row where it belongs into the column, using the column
        #number to print into the correct column
        for j in range(2, len(hoursLog)+2):
            if(columnNo == 1):
                reportSheetCellR = reportSheet.cell(row = j, column = i)
                reportSheetCellR.value = employeeId[zeroIndexFixNested]
                zeroIndexFixNested +=1
                #goes to the next column when length of the column is reached
                if(zeroIndexFixNested == len(hoursLog) and columnNo != len(reportColumnNames)):
                    columnNo+=1
            elif(columnNo == 2):
                reportSheetCellR = reportSheet.cell(row = j, column = i)
                reportSheetCellR.value = nameLog[zeroIndexFixNested]
                zeroIndexFixNested +=1
                #goes to the next column when length of the column is reached
                if(zeroIndexFixNested == len(hoursLog) and columnNo != len(reportColumnNames)):
                    columnNo+=1
            elif(columnNo == 3):
                reportSheetCellR = reportSheet.cell(row = j, column = i)
                reportSheetCellR.value = idLog[zeroIndexFixNested]
                zeroIndexFixNested +=1
                #goes to the next column when length of the column is reached
                if(zeroIndexFixNested == len(hoursLog) and columnNo != len(reportColumnNames)):
                    columnNo+=1
            elif(columnNo == 4):
                reportSheetCellR = reportSheet.cell(row = j, column = i)
                reportSheetCellR.value = hoursLog[zeroIndexFixNested]
                zeroIndexFixNested +=1
                #goes to the next column when length of the column is reached
                if(zeroIndexFixNested == len(hoursLog) and columnNo != len(reportColumnNames)):
                    columnNo+=1
            elif(columnNo == 5):
                reportSheetCellR = reportSheet.cell(row = j, column = i)
                reportSheetCellR.value = performanceStatus[zeroIndexFixNested]
                zeroIndexFixNested +=1
                #goes to the next column when length of the column is reached
                if(zeroIndexFixNested == len(hoursLog) and columnNo != len(reportColumnNames)):
                    columnNo+=1
    #saves to the spreadsheet
    workReport.save("performanceReport.xlsx")
    #ends the count so it can be used to see time elapsed
    global endTime
    endTime = time.time()
    return "Completed"
def diagnosticReport():
    #creates and opens txt so it can be written in, and overrides it if already exists
    diagReport = open("diagnosticReport.txt","w")
    diagReport.write("")
    diagReport.close()
    #opens it to have text added on to it
    diagReport = open("diagnosticReport.txt","a")
    #grabs current date and time
    currentDate = datetime.datetime.now()
    #the generated format for the diagnostic report
    diagReport.write("--------\tDiagnostic Report:\t--------")
    diagReport.write("\nDate:\t"+str(currentDate))
    diagReport.write("\nProject Scan: "+projectScanner())
    diagReport.write("\nEmployee List Scan: "+employeeListScanner())
    diagReport.write("\nPerformance Report Scan: "+performanceReport())
    diagReport.write("\nUnused Elements Scanned: "+str(unusedElements+unusedElementsTwo))
    diagReport.write("\nRuntime(secs): "+str(endTime - startTime))
    diagReport.write("\n--------\tEND\t--------")
    diagReport.close()
    print("Program Complete, spreadsheet made and diagnostic report is available")
diagnosticReport();

